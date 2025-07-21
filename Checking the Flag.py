from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font

import time

# -------------------------------------------------------------------------

# initialize the object
PPT_Robot = webdriver.Chrome()
PPT_Robot.maximize_window()
PPT_Robot.get(" Web App Link ")
hold = WebDriverWait(PPT_Robot,35)




# Extra Mile
def invisibility_overlay():
    """Wait till overlay disappear when navigate"""
    return hold.until(EC.invisibility_of_element_located((By.ID, "overLayDiv" )))


def scrolling():
    """Control Scroll for all"""
    PPT_Robot.execute_script("window.scrollTo(0,450);")


def login_page(username,password) :
    """Login FX"""
    # Check Point To Move.
    hold.until(EC.visibility_of_element_located((By.XPATH,"//div[@id='footer']//p")))

    # Enter UserName
    hold.until(EC.visibility_of_element_located((By.ID,"username"))).send_keys(username)
    # Enter Password
    hold.until(EC.visibility_of_element_located((By.ID, "password"))).send_keys(password)
    # Press Ok To log in
    hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "Button Standard"))).click()
    # Press Customers Button
    hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Customers"))).click()



def ppt_cycle(account_num):
    """Start From Source Page till check ppt flag availability"""
    global done_status
    try :
        def page_source(acc_num):
            """Enter account number to pass source page"""

            # Press Search to view source fields
            hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()

            # Check New Page Elements visibility
            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='id']//p")))

            # select from Drop Menu
            drop_menu = hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "Select Drop Menu")))
            choice = Select(drop_menu)
            choice.select_by_index(0)

            # Enter The Account Number
            hold.until(EC.visibility_of_element_located((By.ID, "acc_num"))).send_keys(acc_num)

            # Press Search.
            hold.until(EC.element_to_be_clickable((By.ID, "ID Name"))).click()
            time.sleep(0.75)
            # Pres to enter the account
            hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "CLASS_NAME"))).click()

        # Run Fx of source page
        page_source(account_num)
        invisibility_overlay()


        def home_page():
            global count_dials,rate_plan,activation_date,done_status,flag_status,status,status_reason

            """fx to check dial ppt flag for each dial"""
            # Check Point To Move Using Footer
            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='id']//p")))
            # Check Point To Move Using Promotions Section
            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='class']//div[@id='id']//span")))

            def re_fetch() :
                """Re-Fetching Home Page For Each Call"""
                # Check Point To Move / Validation For All Account Dials Before Go
                return hold.until(EC.visibility_of_all_elements_located((By.XPATH, "//table[@id='id']//tbody//tr")))
            re_fetch()


            count_dials = len(re_fetch())

            if len(re_fetch()) == 1 :

                rate_plan = re_fetch()[0].find_elements(By.TAG_NAME, "td")[3].text

                if not any(rate in rate_plan for rate in ["Fixed","Gated","eTV"]):

                    for one_dial in range(len(re_fetch())) :
                        re_fetch()

                        activation_date = re_fetch()[one_dial].find_elements(By.TAG_NAME, "td")[5].text

                        scrolling()
                        time.sleep(0.25)

                        # move_to_inner_page() :
                        WebDriverWait(re_fetch()[one_dial],20).until(EC.element_to_be_clickable((By.XPATH,".//td[3]//a[contains(@class,'class')]"))).click()

                        def catch_ppt_flag() :
                            global flag_status
                            # status,status_reason
                            # Check Point using footer
                            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='footer']//p")))
                            # Check Point using invisibility overlay
                            invisibility_overlay()
                            # Check Point using visibility of all table elements
                            hold.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@id='SERVICES TREE']//tbody//tr")))

                            try:
                                list_service_package = hold.until(EC.visibility_of_all_elements_located((By.XPATH, f"//div[@id='SERVICES TABLE']//div[@class='DATreeBody']//table//tbody//tr[@class='DATreeTRE']")))
                                index_list = 0
                                for package in list_service_package:
                                    index_list += 1
                                    if "suppplementary" in package.text.lower() :
                                        hold.until(EC.element_to_be_clickable((By.XPATH, f"//div[@id='SERVICES_TABLE']//div[@class='DATreeBody']//table//tbody//tr[@class='DATreeTRE'][{index_list}]//td[1]//table//tbody//tr//td[1]//a[2]"))).click()
                                        # hold.until(EC.element_to_be_clickable((By.XPATH , f"//td[@class='DATreeItemCell']//a[@name='TreeRowAnchor{}']/following-sibling::a[@class='DATreeItemLink'][1]"))).click()
                                time.sleep(0.5)
                                scrolling()

                                try:
                                    ppt_flag = WebDriverWait(PPT_Robot,2).until(EC.visibility_of_element_located((By.XPATH, "//a[contains(text() , 'Flag PPT dials')]")))
                                    if ppt_flag:
                                        flag_status = "Founded"
                                        PPT_Robot.save_screenshot(f"PPT FLAG{account_num}.png")

                                except:
                                    flag_status = "Not Found"

                            except :
                                flag_status = "supplement field not found"



                        catch_ppt_flag()
                        time.sleep(0.5)



                if any(rate in rate_plan for rate in ["Fixed", "Gated", "eTV"]):
                    flag_status = "null"
                    WebDriverWait(re_fetch()[0], 20).until(EC.element_to_be_clickable((By.XPATH, ".//td[3]//a[contains(@class,'DATblTDALinkTxt')]"))).click()
                    time.sleep(0.5)
                    status_elements = hold.until(EC.visibility_of_all_elements_located((By.XPATH, "//span[@class='DAReadOnlyTxt']")))
                    status = status_elements[0].text
                    status_reason = status_elements[1].text

                    time.sleep(0.5)



            elif len(re_fetch()) != 1:
                flag_status = "Above 1"

        # FX Of Home Page
        home_page()


    except Exception as e :
        done_status = "Not Done"
        PPT_Robot.save_screenshot(f"PPT FLAG{account_num}.png")
        print(f"Issued Account # {account_num}")
        print(f"Error : {e}")
        invisibility_overlay()
        point_zero = hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search")))
        point_zero.click()

    else:
        done_status = "Done"


    finally:
        try :
            invisibility_overlay()
            point_zero = hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search")))
            point_zero.click()
            invisibility_overlay()
            time.sleep(0.5)
        except:

            PPT_Robot.execute_script("history.go(0)")
            invisibility_overlay()
            point_zero = hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search")))
            point_zero.click()
            invisibility_overlay()
            time.sleep(0.5)





# Start Engine Sequence
# ----------------------------------------------------------------------------
print("Opening Sheet")
one_shoot = r" File Path .xlsx"
workbook = excel.load_workbook(one_shoot)
worksheet = workbook["new cycle"]
print("Sheet Now is Opened")
# ----------------------------------------------------------------------------
login_page(" User Name "," Password")



for cell in  range( 2 , worksheet.max_row+1 ) :



    flag_status = ""
    count_dials = 0
    rate_plan = ""
    activation_date = ""
    done_status = ""
    status = ""
    status_reason = ""


    account_number = str(worksheet.cell(cell, 1).value)
    ppt_cycle(account_number)

    print(f"{cell-1} : Account || {account_number} in progress....")



    worksheet.cell(cell, 2 ).value = flag_status
    #
    # worksheet.cell(cell, 7 ).value = count_dials
    #
    # worksheet.cell(cell, 8 ).value = rate_plan
    #
    # worksheet.cell(cell, 9 ).value = activation_date
    #
    # worksheet.cell(cell, 10).value = done_status
    #
    # worksheet.cell(cell, 11).value = status
    #
    # worksheet.cell(cell, 12).value = status_reason


    workbook.save(one_shoot)




PPT_Robot.quit()